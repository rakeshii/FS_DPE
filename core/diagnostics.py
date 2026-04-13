"""
core/diagnostics.py
DiagnosticsReport — accumulates per-sheet processing records and
global warnings, then renders them into an 'Update Report' sheet
in the output workbook.

Interface is driven entirely by how projector.py uses this module:

    diag = DiagnosticsReport(new_header=new_header)
    diag.global_warnings.append(str(issue))
    diag.add(SheetRecord(
        sheet_name, sheet_type, config_insert_col, detected_insert_col,
        detection_confidence, current_year, new_year_col, frozen_year_col,
        formulas_copied, missing_cache_cells=0, missing_cache_refs=[],
        warnings=[]
    ))
    create_update_report_sheet(wb, diag)
"""

from dataclasses import dataclass, field
from typing import List, Optional
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class SheetRecord:
    sheet_name          : str
    sheet_type          : str           # 'standard' | 'wide' | 'fixed_assets'
    config_insert_col   : int           # value from config.py
    detected_insert_col : int           # value from column_detector
    detection_confidence: float         # 0.0 – 1.0
    current_year        : int           # e.g. 2025  (0 = unknown)
    new_year_col        : int           # 1-indexed col where 2026 was written
    frozen_year_col     : int           # 1-indexed col that was frozen
    formulas_copied     : int           # number of formulas copied to 2026 col
    missing_cache_cells : int = 0       # formula cells with no cached value
    missing_cache_refs  : List[str] = field(default_factory=list)
    warnings            : List[str] = field(default_factory=list)

    @property
    def col_matched(self) -> bool:
        return self.config_insert_col == self.detected_insert_col

    @property
    def confidence_label(self) -> str:
        c = self.detection_confidence
        if c >= 0.80: return 'HIGH'
        if c >= 0.50: return 'MEDIUM'
        if c >= 0.30: return 'LOW'
        return 'VERY LOW'


@dataclass
class DiagnosticsReport:
    new_header     : str
    records        : List[SheetRecord] = field(default_factory=list)
    global_warnings: List[str]         = field(default_factory=list)

    def add(self, record: SheetRecord):
        self.records.append(record)

    @property
    def total_formulas(self) -> int:
        return sum(r.formulas_copied for r in self.records)

    @property
    def total_missing(self) -> int:
        return sum(r.missing_cache_cells for r in self.records)

    @property
    def mismatched_sheets(self) -> List[SheetRecord]:
        return [r for r in self.records if not r.col_matched]


# ── Report sheet renderer ─────────────────────────────────────────────────────

def create_update_report_sheet(wb, diag: DiagnosticsReport):
    """
    Append an 'Update Report' sheet to `wb` documenting every
    column-detection and formula-migration decision made during this run.
    """
    SHEET_NAME = 'Update Report'
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    # ── Colour palette ────────────────────────────────────────────────
    BRAND      = '0585A8'
    BRAND_DK   = '045F78'
    LIGHT_BLUE = 'D6EFF5'
    GREEN_BG   = 'EAF6EA'
    GREEN_DK   = '1B5E20'
    AMBER_BG   = 'FFF9E6'
    AMBER_DK   = '7A5200'
    RED_BG     = 'FDECEA'
    RED_DK     = 'B71C1C'
    GRAY       = 'F5F6F8'
    WHITE      = 'FFFFFF'

    thin  = Side(style='thin',   color='C8D8E4')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _s(cell, bg, fg, bold=False, size=9, wrap=False, align='left'):
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.font      = Font(bold=bold, size=size, color=fg, name='Calibri')
        cell.alignment = Alignment(horizontal=align, vertical='center',
                                   wrap_text=wrap)
        cell.border    = brd

    def _merged(row, text, bg, fg, bold=True, size=10, end_col=8):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=end_col)
        c = ws.cell(row, 1, text)
        _s(c, bg, fg, bold=bold, size=size, wrap=True, align='left')
        ws.row_dimensions[row].height = 18

    # ── Column widths ─────────────────────────────────────────────────
    col_widths = [22, 12, 10, 10, 14, 12, 12, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.merge_cells(f'A{row}:H{row}')
    c = ws.cell(row, 1, f'Update Report  —  {diag.new_header}')
    _s(c, BRAND, WHITE, bold=True, size=12, align='center')
    ws.row_dimensions[row].height = 22
    row += 1

    # Summary line
    ws.merge_cells(f'A{row}:H{row}')
    summary = (
        f"Sheets processed: {len(diag.records)}  |  "
        f"Formulas copied: {diag.total_formulas}  |  "
        f"Missing cache cells: {diag.total_missing}  |  "
        f"Col-config mismatches: {len(diag.mismatched_sheets)}"
    )
    c = ws.cell(row, 1, summary)
    _s(c, LIGHT_BLUE, BRAND_DK, bold=False, size=9, wrap=True, align='left')
    ws.row_dimensions[row].height = 16
    row += 1

    # Global warnings
    if diag.global_warnings:
        row += 1
        _merged(row, 'GLOBAL WARNINGS', AMBER_BG, AMBER_DK, size=9)
        row += 1
        for warn in diag.global_warnings:
            ws.merge_cells(f'A{row}:H{row}')
            c = ws.cell(row, 1, f'  ⚠  {warn}')
            _s(c, AMBER_BG, AMBER_DK, size=9, wrap=True)
            ws.row_dimensions[row].height = 30
            row += 1

    # Per-sheet table
    row += 1
    _merged(row, 'PER-SHEET PROCESSING DETAIL', BRAND_DK, WHITE, size=9)
    row += 1

    # Table header
    headers = ['Sheet', 'Type', 'Cfg col', 'Det col', 'Confidence',
               'Formulas', 'Missing', 'Warnings / Notes']
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row, col, h)
        _s(c, BRAND, WHITE, bold=True, size=9, align='center')
    row += 1

    for rec in diag.records:
        col_match = rec.col_matched
        conf      = rec.detection_confidence

        # Row background based on outcome
        if rec.warnings or not col_match or rec.missing_cache_cells > 0:
            bg, fg = AMBER_BG, AMBER_DK
        else:
            bg, fg = GREEN_BG, GREEN_DK

        cfg_ltr = get_column_letter(rec.config_insert_col)
        det_ltr = get_column_letter(rec.detected_insert_col)

        note_parts = list(rec.warnings)
        if not col_match:
            note_parts.insert(0,
                f'Config col {cfg_ltr} overridden → detected col {det_ltr}')
        if rec.missing_cache_refs:
            note_parts.append(
                f'No cached value for: {", ".join(rec.missing_cache_refs[:5])}'
                + (f' … (+{len(rec.missing_cache_refs)-5} more)'
                   if len(rec.missing_cache_refs) > 5 else ''))
        notes_str = '  |  '.join(note_parts) if note_parts else '—'

        values = [
            rec.sheet_name,
            rec.sheet_type,
            cfg_ltr,
            det_ltr,
            f'{conf:.0%}  ({rec.confidence_label})',
            rec.formulas_copied,
            rec.missing_cache_cells,
            notes_str,
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row, col, val)
            _s(c, bg, fg, size=9,
               wrap=(col == 8),
               align='center' if col in (2, 3, 4, 5, 6, 7) else 'left')
        ws.row_dimensions[row].height = 15
        row += 1

    # Legend
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    c = ws.cell(row, 1,
        'Legend:  Green = clean run  |  '
        'Amber = config override, missing cache, or warnings present  |  '
        'Cfg col = column from config.py  |  Det col = column from scan')
    _s(c, LIGHT_BLUE, BRAND_DK, size=9, wrap=True)
    ws.row_dimensions[row].height = 18

    ws.freeze_panes = 'A6'