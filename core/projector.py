"""
core/projector.py
Pure v8 business logic - no Flask, no HTTP.
Exact same formula engine as the verified FS-FY_2025-26_v8_FINAL.xlsx.
"""

import os, re, copy, subprocess, tempfile
from openpyxl.workbook.defined_name import DefinedName
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment

# ── Compiled regex (v8 cross-sheet range aware) ─────────────────
FULL_RANGE_PAT = re.compile(
    r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9_]*))!"
    r"(\$?)([A-Z]+)(\$?\d+)"
    r"(?::(\$?)([A-Z]+)(\$?\d+))?"
)
SAME_PAT  = re.compile(r"(\$?)([A-Z]+)(\$?\d+)")
PERIOD_RE = re.compile(r'31 March|March 31|Year Ended|31-03-|March, 202', re.I)

# Support sheets - their refs are fixed source data, never shifted for 2026
SUPPORT_SHEETS = frozenset({
    'TB', 'WDV Dep-IT', 'RATIO WORKING', 'RATIO SUMMARY',
    'COMPUTATION', 'CSR', 'Related Party info'
})


# ══════════════════════════════════════════════════════════════
# FREEZE 2025 COLUMN - blank all data, keep only period header labels
# ══════════════════════════════════════════════════════════════

def freeze_2025_col_with_values(ws, col_2025, cached_values):
    """
    Replace every cell in the 2025 column with its pre-computed value.
    - Formula cells  → replaced with their cached computed result (plain number)
    - Plain number cells → kept as-is
    - Period header labels → kept as-is
    - None/empty cells → kept as None
    This gives auditors the 2025 actual figures as plain numbers (no formulas),
    exactly matching the pattern of the 2024 column in the original file.
    """
    for r in range(1, ws.max_row + 1):
        cell    = ws.cell(r, col_2025)
        current = cell.value           # this is the shifted 2025 formula after insert
        cached  = cached_values.get(r) # pre-computed result from data_only read

        if cached is None:
            # No cached value - keep whatever is there (header text, None)
            if isinstance(current, str) and current.startswith('='):
                cell.value = None      # formula with no cached result → blank
            continue

        if isinstance(cached, str) and PERIOD_RE.search(cached):
            continue   # period header label - keep as-is

        # Write the computed value (number or text result)
        cell.value = cached


# ══════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════

def generate_projection(input_path, output_path, sheet_config, chain_patches,
                        new_header='As at 31 March, 2026', copy_vals=True,
                        update_titles=True, lo_cmd='libreoffice', lo_timeout=90,
                        named_ranges=None):
    xlsx_path = _ensure_xlsx(input_path, lo_cmd, lo_timeout)
    try:
        # Step 1: Read cached formula results (data_only) BEFORE modifying anything
        # These become the frozen 2025 values in the output
        wb_data = openpyxl.load_workbook(xlsx_path, keep_links=False, data_only=True)
        cached = {}   # { shname: { row: computed_value } }
        for shname, cfg in sheet_config.items():
            if shname not in wb_data.sheetnames:
                continue
            ws_d = wb_data[shname]
            col   = cfg['insert']
            width = cfg.get('width', 1)
            cached[shname] = {}
            for r in range(1, ws_d.max_row + 1):
                cached[shname][r] = ws_d.cell(r, col).value
            if width == 2:
                # Wide sheet: cache second sub-col (₹ amount = col+1) separately
                cached[shname + '_col2'] = {}
                for r in range(1, ws_d.max_row + 1):
                    cached[shname + '_col2'][r] = ws_d.cell(r, col + 1).value
        wb_data.close()

        # Step 2: Load for editing, process all sheets
        wb = openpyxl.load_workbook(xlsx_path, keep_links=False)
        for shname in wb.sheetnames:
            cfg = sheet_config.get(shname)
            if cfg:
                width = cfg.get('width', 1)
                if width == 2:
                    # Wide sheet: insert 2 cols, 2026 gets No.of Shares + ₹ formula cols
                    process_wide_financial_sheet(wb[shname], shname, cfg['insert'],
                                                 sheet_config, new_header, update_titles)
                    # Freeze 2025 sub-cols (now shifted right by 2)
                    freeze_2025_col_with_values(wb[shname], cfg['insert'] + 2,
                                                cached.get(shname, {}))
                    freeze_2025_col_with_values(wb[shname], cfg['insert'] + 3,
                                                cached.get(shname + '_col2', {}))
                else:
                    process_financial_sheet(wb[shname], shname, cfg['insert'],
                                            sheet_config, new_header, copy_vals, update_titles)
                    freeze_2025_col_with_values(wb[shname], cfg['insert'] + 1,
                                                cached.get(shname, {}))
            else:
                process_support_sheet(wb[shname], sheet_config)
        _apply_chain_patches(wb, sheet_config, chain_patches)

        # Create Named Ranges in the workbook for all TB row references
        if named_ranges:
            create_tb_named_ranges(wb, named_ranges)
            rewrite_tb_formulas(wb, named_ranges, sheet_config)

        # Create year-end checklist sheet
        create_checklist_sheet(wb, sheet_config, named_ranges)
        wb.save(output_path)
    finally:
        if xlsx_path != input_path and os.path.exists(xlsx_path):
            try: os.remove(xlsx_path)
            except OSError: pass



# ══════════════════════════════════════════════════════════════
# YEAR-END CHECKLIST SHEET
# ══════════════════════════════════════════════════════════════

def create_checklist_sheet(wb, sheet_config, named_ranges_cfg=None):
    """
    Creates a "Year-End Checklist" sheet in the output workbook.
    Contains:
    1. Instructions for the accountant
    2. Snapshot of all TB accounts with their coverage status
    3. Named range block boundaries (to check if new rows fell outside)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    SHEET_NAME = 'Year-End Checklist'
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)

    # ── Colours ────────────────────────────────────────────────
    BRAND      = '0585A8'
    BRAND_DK   = '045F78'
    LIGHT_BLUE = 'D6EFF5'
    AMBER_BG   = 'FFF9E6'
    AMBER_DK   = '7A5200'
    GREEN_BG   = 'EAF6EA'
    GREEN_DK   = '1B5E20'
    RED_BG     = 'FDECEA'
    RED_DK     = 'B71C1C'
    GRAY       = 'F5F6F8'
    WHITE      = 'FFFFFF'

    thin = Side(style='thin', color='C8D8E4')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(cell, bg, fg, bold=False, size=10, wrap=False, align='left'):
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.font      = Font(bold=bold, size=size, color=fg,
                              name='Calibri')
        cell.alignment = Alignment(horizontal=align, vertical='center',
                                   wrap_text=wrap)
        cell.border    = brd

    def header_row(row, texts, widths, bg=BRAND, fg=WHITE, bold=True):
        for col, (text, w) in enumerate(zip(texts, widths), start=1):
            c = ws.cell(row, col, text)
            style(c, bg, fg, bold=bold, size=10, align='center')
            ws.column_dimensions[get_column_letter(col)].width = w

    def merged_header(row, text, bg, fg, bold=True, size=11):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=6)
        c = ws.cell(row, 1, text)
        style(c, bg, fg, bold=bold, size=size, wrap=True, align='left')
        ws.row_dimensions[row].height = 20

    # ── Column widths ──────────────────────────────────────────
    COL_WIDTHS = [8, 42, 16, 16, 22, 28]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Title ──────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:F{row}')
    c = ws.cell(row, 1, 'Year-End Checklist  -  FinStatement Projector')
    style(c, BRAND, WHITE, bold=True, size=13, align='center')
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(f'A{row}:F{row}')
    c = ws.cell(row, 1,
        'Use this sheet every year before sign-off. '
        'Compare new TB accounts against the list below and verify coverage.')
    style(c, LIGHT_BLUE, BRAND_DK, bold=False, size=10, wrap=True, align='left')
    ws.row_dimensions[row].height = 18
    row += 1

    # ── Step-by-step instructions ──────────────────────────────
    row += 1
    merged_header(row, 'STEP-BY-STEP INSTRUCTIONS', BRAND_DK, WHITE, size=10)
    row += 1

    steps = [
        ('Step 1', 'Open the new year TB sheet. Press Ctrl+End - note the last row with data.'),
        ('Step 2', 'Compare against column A below. Any row that did NOT exist last year = new account.'),
        ('Step 3', 'For each new account - check column E below. Is it covered by a named range?'),
        ('Step 4', 'If a new account falls OUTSIDE a range block (col F) - extend the named range. Formulas > Name Manager > select range > Edit > expand boundary.'),
        ('Step 5', 'If a new account is a BRAND NEW category not in any Note sheet - add it manually to the relevant Note sheet and create a new Named Range for it.'),
        ('Step 6', 'If an account was DELETED from TB - find the #REF! error in the FS and remove it.'),
    ]
    for step, desc in steps:
        ws.merge_cells(f'A{row}:F{row}')
        c = ws.cell(row, 1, f'  {step}: {desc}')
        style(c, AMBER_BG, AMBER_DK, bold=False, size=9, wrap=True, align='left')
        ws.row_dimensions[row].height = 30 if '\n' in desc else 18
        row += 1

    # ── Named range blocks ─────────────────────────────────────
    row += 1
    merged_header(row, 'NAMED RANGE BLOCKS  -  Critical boundaries to check', BRAND_DK, WHITE, size=10)
    row += 1

    header_row(row,
               ['Block name', 'TB rows covered', 'Debit col', 'Credit col',
                'Action if new row added outside boundary'],
               [25, 20, 15, 15, 45],
               bg=BRAND, fg=WHITE)
    row += 1

    blocks = [
        ('All salaries',         '126 – 169', 'TB_B_All_Salaries_B',         'TB_C_All_Salaries_C',
         'Extend range to include new row. Name Manager → TB_B_All_Salaries_B → Edit'),
        ('TDS (48-54)',          '48 – 54',   'TB_B_TDS_Block_48_54_B',      'TB_C_TDS_Block_48_54_C',
         'Extend both B and C ranges'),
        ('TCS (45-47)',          '45 – 47',   'TB_B_TDS_Block_45_47_B',      'TB_C_TDS_Block_45_47_C',
         'Extend both B and C ranges'),
        ('Purchase types',       '102 – 106', 'TB_B_Purchase_Block_102_106_B','TB_C_Purchase_Block_102_106_C',
         'Extend both B and C ranges'),
        ('CSR expenses',         '119 – 120', 'TB_B_CSR_Block_119_120_B',    'TB_C_CSR_Block_119_120_C',
         'Extend both B and C ranges'),
    ]
    for i, (name, rows, nb, nc, action) in enumerate(blocks):
        bg = WHITE if i % 2 == 0 else GRAY
        for col, val in enumerate([name, rows, nb, nc, action], 1):
            c = ws.cell(row, col, val)
            style(c, bg, '333333', bold=False, size=9,
                  wrap=(col == 5), align='left')
        row += 1

    # ── TB Account snapshot ────────────────────────────────────
    row += 1
    merged_header(row, 'TB ACCOUNT SNAPSHOT  -  Baseline at time of generation', BRAND_DK, WHITE, size=10)
    row += 1

    header_row(row,
               ['TB row', 'Account name', 'Debit (₹)', 'Credit (₹)',
                'Coverage', 'Note / Named range'],
               COL_WIDTHS,
               bg=BRAND, fg=WHITE)
    row += 1

    # Gather TB data
    if 'TB' not in wb.sheetnames:
        return

    ws_tb = wb['TB']

    # Build set of referenced rows from named ranges config
    import re as _re
    referenced_rows = set()
    if named_ranges_cfg:
        for slug, (row_or_range, cols) in named_ranges_cfg.items():
            if isinstance(row_or_range, tuple):
                start, end = row_or_range
                for r2 in range(start, end+1):
                    referenced_rows.add(r2)
            else:
                referenced_rows.add(row_or_range)

    # Also scan formulas for direct TB refs
    for shname, cfg in sheet_config.items():
        if shname not in wb.sheetnames: continue
        ws_fs = wb[shname]
        for r2 in range(1, ws_fs.max_row+1):
            for c2 in range(1, ws_fs.max_column+1):
                v = ws_fs.cell(r2, c2).value
                if isinstance(v, str) and 'TB!' in v:
                    for m in _re.finditer(r'TB![$]?[BC][$]?([0-9]+)', v):
                        referenced_rows.add(int(m.group(1)))

    # Name lookup
    name_lookup = {}
    if named_ranges_cfg:
        for slug, (row_or_range, cols) in named_ranges_cfg.items():
            if isinstance(row_or_range, tuple):
                start, end = row_or_range
                name_lookup[(start, end)] = f'TB_B_{slug} (rows {start}-{end})'
            else:
                name_lookup[row_or_range] = f'TB_B_{slug}'

    for r2 in range(8, ws_tb.max_row + 1):
        acct = ws_tb.cell(r2, 1).value
        deb  = ws_tb.cell(r2, 2).value
        crd  = ws_tb.cell(r2, 3).value

        if not acct or not isinstance(acct, str) or len(acct.strip()) < 2:
            continue

        is_header = (deb is None and crd is None)
        has_value = not is_header

        covered   = r2 in referenced_rows
        bg_row    = GREEN_BG if covered else (AMBER_BG if has_value else GRAY)
        fg_row    = GREEN_DK if covered else (AMBER_DK if has_value else '666666')

        # Coverage label
        if covered:
            coverage = 'Named range'
        elif is_header:
            coverage = 'Group header'
        else:
            coverage = 'NOT COVERED'

        # Named range label
        nr_label = name_lookup.get(r2, '')
        if not nr_label and covered:
            # Check if inside a range block
            for key, label in name_lookup.items():
                if not isinstance(key, tuple): continue
                start, end = key
                if True:
                    if start <= r2 <= end:
                        nr_label = label
                        break

        row_data = [
            r2,
            acct.strip(),
            deb if deb else '',
            crd if crd else '',
            coverage,
            nr_label
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row, col, val)
            num_fmt = '#,##0.00' if col in (3, 4) and isinstance(val, (int, float)) else None
            style(c, bg_row, fg_row, bold=False,
                  size=9, wrap=(col in (2, 6)), align='right' if col in (1,3,4) else 'left')
            if num_fmt:
                c.number_format = num_fmt
        ws.row_dimensions[row].height = 15
        row += 1

    # ── Legend ─────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    c = ws.cell(row, 1,
        'Legend:  Green = covered by named range  |  '
        'Amber = has value but NOT covered (verify this account is in correct Note sheet)  |  '
        'Gray = group header (no value, no coverage needed)')
    style(c, LIGHT_BLUE, BRAND_DK, bold=False, size=9, wrap=True, align='left')
    ws.row_dimensions[row].height = 22

    # Freeze panes below instruction rows
    ws.freeze_panes = 'A14'


# ══════════════════════════════════════════════════════════════
# LIBREOFFICE - auto-detect Windows / Mac / Linux
# ══════════════════════════════════════════════════════════════

def _find_libreoffice():
    import shutil, platform, glob
    env_cmd = os.environ.get('LIBREOFFICE_CMD', '').strip()
    if env_cmd and shutil.which(env_cmd):
        return env_cmd
    system = platform.system()
    if system == 'Windows':
        candidates = [
            r'C:\Program Files\LibreOffice\program\soffice.exe',
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
        ]
        candidates += glob.glob(r'C:\Program Files\LibreOffice*\program\soffice.exe')
        candidates += glob.glob(r'C:\Program Files (x86)\LibreOffice*\program\soffice.exe')
        for path in candidates:
            if os.path.isfile(path): return path
    elif system == 'Darwin':
        mac = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
        if os.path.isfile(mac): return mac
    for cmd in ['libreoffice', 'soffice']:
        found = shutil.which(cmd)
        if found: return found
    raise RuntimeError(
        "LibreOffice is not installed or not found.\n\n"
        "Please install it and restart the server:\n"
        "  Windows : https://www.libreoffice.org/download/download-libreoffice/\n"
        "  Mac     : brew install --cask libreoffice\n"
        "  Linux   : sudo apt install libreoffice\n\n"
        "After installing, run:  python app.py"
    )


def _ensure_xlsx(input_path, lo_cmd, timeout):
    if input_path.lower().endswith('.xlsx'):
        return input_path
    lo_exe = _find_libreoffice()
    tmpdir = tempfile.mkdtemp()
    try:
        result = subprocess.run(
            [lo_exe, '--headless', '--convert-to', 'xlsx', '--outdir', tmpdir, input_path],
            capture_output=True, text=True, timeout=timeout
        )
    except FileNotFoundError:
        raise RuntimeError(f"LibreOffice not found at: {lo_exe}\nDownload from https://www.libreoffice.org/download/")
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed (exit {result.returncode}).\n{result.stderr[:400]}")
    base     = os.path.splitext(os.path.basename(input_path))[0]
    out_path = os.path.join(tmpdir, base + '.xlsx')
    if not os.path.exists(out_path):
        raise RuntimeError(f"LibreOffice ran but output .xlsx not found: {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════════
# FORMULA SHIFTER - exact v8 logic
# ══════════════════════════════════════════════════════════════

def shift_formula(formula, cur_insert, sheet_config=None):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    cfg = sheet_config or {}

    def shift_cross(m):
        ref_sh = m.group(1) or m.group(2)
        abs1, col1, row1 = m.group(3), m.group(4), m.group(5)
        abs2, col2, row2 = m.group(6), m.group(7), m.group(8)
        ri  = cfg.get(ref_sh, {}).get('insert')
        pfx = f"'{ref_sh}'" if re.search(r"[ &\-]", ref_sh) else ref_sh
        cn1 = column_index_from_string(col1)
        if ri and cn1 >= ri: col1 = get_column_letter(cn1 + 1)
        if col2:
            cn2 = column_index_from_string(col2)
            if ri and cn2 >= ri: col2 = get_column_letter(cn2 + 1)
            return f"{pfx}!{abs1}{col1}{row1}:{abs2}{col2}{row2}"
        return f"{pfx}!{abs1}{col1}{row1}"

    def shift_same(m):
        abs_c, col_ltr, row_part = m.group(1), m.group(2), m.group(3)
        if cur_insert is None: return m.group(0)
        cn = column_index_from_string(col_ltr)
        if cn >= cur_insert: col_ltr = get_column_letter(cn + 1)
        return f"{abs_c}{col_ltr}{row_part}"

    result, last = [], 0
    for m in FULL_RANGE_PAT.finditer(formula):
        result.append(SAME_PAT.sub(shift_same, formula[last:m.start()]))
        result.append(shift_cross(m))
        last = m.end()
    result.append(SAME_PAT.sub(shift_same, formula[last:]))
    return ''.join(result)


# ══════════════════════════════════════════════════════════════
# DERIVE 2026 FORMULA - shift 2025 formula one column LEFT
# ══════════════════════════════════════════════════════════════

def derive_2026_formula(formula_2025, col_2025_1idx, sheet_config=None):
    """
    Derive the correct 2026 formula from the already-shifted 2025 formula.

    Rule - shift every same-sheet column ref >= insert_col one position LEFT:
      col_2025 refs  →  col_2026  (self-references)
      col_2024 refs  →  col_2025  (backward-looking refs like opening balance)
      any col N >= insert_col  →  N-1

    Cross-sheet refs to financial sheets: shift their 2025 col → their 2026 col
    Cross-sheet refs to support sheets (TB, COMPUTATION etc): unchanged

    Examples:
      2025 col E: =SUM(E9:E12)          → 2026 col D: =SUM(D9:D12)
      2025 col E: =+F13  (F=2024 col)   → 2026 col D: =+E13  (E=2025 col)
      2025 col D: =+E34  (E=2024 col)   → 2026 col C: =+D34  (D=2025 col)
      2025 col D: =+TB!C21              → 2026 col C: =+TB!C21  (TB unchanged)
    """
    if not isinstance(formula_2025, str) or not formula_2025.startswith('='):
        return formula_2025

    cfg        = sheet_config or {}
    insert_col = col_2025_1idx - 1   # insert_col == col_2026_1idx

    def _shift_cross(m):
        ref_sh = m.group(1) or m.group(2)
        abs1, col1, row1 = m.group(3), m.group(4), m.group(5)
        abs2, col2, row2 = m.group(6), m.group(7), m.group(8)
        pfx = f"'{ref_sh}'" if re.search(r"[ &\-]", ref_sh) else ref_sh

        if ref_sh in SUPPORT_SHEETS:
            if col2: return f"{pfx}!{abs1}{col1}{row1}:{abs2}{col2}{row2}"
            return f"{pfx}!{abs1}{col1}{row1}"

        ref_ins = cfg.get(ref_sh, {}).get('insert')
        if ref_ins:
            # In the referenced sheet: 2025 col = ref_ins+1, 2026 col = ref_ins
            ref_2025 = ref_ins + 1
            ref_2026 = ref_ins
            cn1 = column_index_from_string(col1)
            if cn1 == ref_2025: col1 = get_column_letter(ref_2026)
            if col2:
                cn2 = column_index_from_string(col2)
                if cn2 == ref_2025: col2 = get_column_letter(ref_2026)
                return f"{pfx}!{abs1}{col1}{row1}:{abs2}{col2}{row2}"
            return f"{pfx}!{abs1}{col1}{row1}"

        if col2: return f"{pfx}!{abs1}{col1}{row1}:{abs2}{col2}{row2}"
        return f"{pfx}!{abs1}{col1}{row1}"

    def _shift_same(m):
        abs_c, col_ltr, row_part = m.group(1), m.group(2), m.group(3)
        cn = column_index_from_string(col_ltr)
        # Shift ALL cols >= insert_col one step LEFT
        # This correctly handles:
        #   col_2025 → col_2026  (self-references)
        #   col_2024 → col_2025  (backward-looking: opening = prev year closing)
        #   any further-right col also shifts left by 1
        if cn >= insert_col:
            col_ltr = get_column_letter(cn - 1)
        return f"{abs_c}{col_ltr}{row_part}"

    result, last = [], 0
    for m in FULL_RANGE_PAT.finditer(formula_2025):
        result.append(SAME_PAT.sub(_shift_same, formula_2025[last:m.start()]))
        result.append(_shift_cross(m))
        last = m.end()
    result.append(SAME_PAT.sub(_shift_same, formula_2025[last:]))
    return ''.join(result)


# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════

def copy_style(src, dst):
    if src.has_style:
        dst.font = copy.copy(src.font); dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border); dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format


def is_period_hdr(ws, r, max_col):
    for c in range(1, max_col + 2):
        if PERIOD_RE.search(str(ws.cell(r, c).value or '')): return True
    return False


# ══════════════════════════════════════════════════════════════
# PROCESS FINANCIAL SHEET - exact v8 logic
# ══════════════════════════════════════════════════════════════

def process_financial_sheet(ws, shname, insert_col, sheet_config,
                             new_header, copy_vals=True, update_titles=True):
    max_row, max_col = ws.max_row, ws.max_column

    # Save merges
    saved_merges = []
    for mr in list(ws.merged_cells.ranges):
        saved_merges.append({'min_row':mr.min_row,'max_row':mr.max_row,
                             'min_col':mr.min_col,'max_col':mr.max_col,
                             'val':ws.cell(mr.min_row,mr.min_col).value})
    for mr in list(ws.merged_cells.ranges): ws.unmerge_cells(str(mr))

    # Save col dims
    orig_dims = {l:(d.width,d.hidden) for l,d in ws.column_dimensions.items()}

    # Shift cells right-to-left
    for c in range(max_col, insert_col - 1, -1):
        for r in range(1, max_row + 1):
            src = ws.cell(r, c); dst = ws.cell(r, c + 1)
            val = src.value
            if isinstance(val, str) and val.startswith('='):
                val = shift_formula(val, insert_col, sheet_config)
            dst.value = val
            copy_style(src, dst)

    # Clear insert column
    for r in range(1, max_row + 1): ws.cell(r, insert_col).value = None

    # Re-apply merges (shifted)
    for m in saved_merges:
        sc = lambda c: c + 1 if c >= insert_col else c
        ws.merge_cells(start_row=m['min_row'], start_column=sc(m['min_col']),
                       end_row=m['max_row'],   end_column=sc(m['max_col']))
        if m['val'] is not None:
            ws.cell(m['min_row'], sc(m['min_col'])).value = m['val']

    # Restore col dims (shifted)
    for letter,(width,hidden) in orig_dims.items():
        oc = column_index_from_string(letter)
        nc = oc + 1 if oc >= insert_col else oc
        ws.column_dimensions[get_column_letter(nc)].width  = width
        ws.column_dimensions[get_column_letter(nc)].hidden = hidden
    ws.column_dimensions[get_column_letter(insert_col)].width = 22

    # Update title dates
    if update_titles:
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    if '2024-25' in v:
                        ws.cell(r, c).value = v.replace('2024-25', '2025-26')
                    elif re.search(r'31st? March,? 2025', v):
                        ws.cell(r, c).value = v.replace('2025', '2026')

    # Write 2026 column
    col_2025 = insert_col + 1
    for r in range(1, max_row + 1):
        c26 = ws.cell(r, insert_col)
        c25 = ws.cell(r, col_2025)
        v25 = c25.value

        if is_period_hdr(ws, r, max_col):
            v_str = str(v25 or '')
            if any(k in v_str for k in ['2025','31 March','March 31','Year Ended']):
                c26.value     = new_header
                c26.font      = Font(bold=True, size=9, color="7A5200")
                c26.fill      = PatternFill("solid", fgColor="FFF0C0")
                c26.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif isinstance(v25, str) and v25.startswith('='):
            # 2025 had a formula → derive the matching 2026 formula (shifted one col left)
            c26.value = derive_2026_formula(v25, col_2025, sheet_config)
            copy_style(c25, c26)
            c26.fill = PatternFill("solid", fgColor="FFFBEB")
            c26.font = Font(size=9, color="856404")
        # Plain numbers from 2025 are NOT copied to 2026.
        # 2026 col follows the same pattern as the current year (2025):
        # only formulas are present; auditor fills TB → values auto-populate.


# ══════════════════════════════════════════════════════════════
# PROCESS WIDE FINANCIAL SHEET - two sub-cols per year (e.g. Note 2)
# ══════════════════════════════════════════════════════════════

def process_wide_financial_sheet(ws, shname, insert_col, sheet_config,
                                  new_header, update_titles=True):
    """
    Handle sheets where each year has TWO sub-columns (e.g. Note 2: No.of Shares + ₹).
    Inserts TWO new columns at insert_col:
      insert_col   = 2026 No. of Shares  (blank - auditor fills)
      insert_col+1 = 2026 ₹ amount       (formulas derived from 2025 ₹ col)
    The 2025 sub-cols shift right by 2.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # Save merged ranges
    saved_merges = []
    for mr in list(ws.merged_cells.ranges):
        saved_merges.append({
            'min_row': mr.min_row, 'max_row': mr.max_row,
            'min_col': mr.min_col, 'max_col': mr.max_col,
            'val': ws.cell(mr.min_row, mr.min_col).value,
        })
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # Save column dimensions
    orig_dims = {l: (d.width, d.hidden) for l, d in ws.column_dimensions.items()}

    # Shift all cells right by 2 (from insert_col onwards)
    for c in range(max_col, insert_col - 1, -1):
        for r in range(1, max_row + 1):
            src = ws.cell(r, c)
            dst = ws.cell(r, c + 2)
            val = src.value
            if isinstance(val, str) and val.startswith('='):
                val = shift_formula(val, insert_col, sheet_config)
                # shift again for the second inserted col
                val = shift_formula(val, insert_col + 1, sheet_config)
            dst.value = val
            copy_style(src, dst)

    # Clear the two new cols
    for r in range(1, max_row + 1):
        ws.cell(r, insert_col).value     = None
        ws.cell(r, insert_col + 1).value = None

    # Re-apply merged cells (shift by 2)
    for m in saved_merges:
        sc = lambda c: c + 2 if c >= insert_col else c
        ws.merge_cells(
            start_row=m['min_row'], start_column=sc(m['min_col']),
            end_row=m['max_row'],   end_column=sc(m['max_col'])
        )
        if m['val'] is not None:
            ws.cell(m['min_row'], sc(m['min_col'])).value = m['val']

    # Restore column dimensions (shift by 2)
    for letter, (width, hidden) in orig_dims.items():
        oc = column_index_from_string(letter)
        nc = oc + 2 if oc >= insert_col else oc
        ws.column_dimensions[get_column_letter(nc)].width  = width
        ws.column_dimensions[get_column_letter(nc)].hidden = hidden
    ws.column_dimensions[get_column_letter(insert_col)].width     = 18
    ws.column_dimensions[get_column_letter(insert_col + 1)].width = 22

    # Optional: update title dates
    if update_titles:
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    if '2024-25' in v:
                        ws.cell(r, c).value = v.replace('2024-25', '2025-26')
                    elif re.search(r'31st? March,? 2025', v):
                        ws.cell(r, c).value = v.replace('2025', '2026')

    # Identify what's in the original 2025 sub-cols (now at insert_col+2 and insert_col+3)
    col_2025_shares = insert_col + 2   # No. of Shares col (was insert_col in original)
    col_2025_amount = insert_col + 3   # ₹ amount col      (was insert_col+1 in original)
    col_2026_shares = insert_col       # new: 2026 No. of Shares
    col_2026_amount = insert_col + 1   # new: 2026 ₹ amount

    for r in range(1, max_row + 1):
        v_shares = ws.cell(r, col_2025_shares).value
        v_amount = ws.cell(r, col_2025_amount).value

        # Write period header in the 2026 header cell
        if is_period_hdr(ws, r, max_col + 2):
            v_str = str(v_shares or '') + str(v_amount or '')
            if any(k in v_str for k in ['2025', '31 March', 'March 31', 'Year Ended']):
                ws.cell(r, col_2026_shares).value     = new_header
                ws.cell(r, col_2026_shares).font      = Font(bold=True, size=9, color="7A5200")
                ws.cell(r, col_2026_shares).fill      = PatternFill("solid", fgColor="FFF0C0")
                ws.cell(r, col_2026_shares).alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
            continue

        # Mirror sub-column headers (No. of Shares / ₹ / % etc.)
        # These are non-formula text labels that sit one row below the period header
        if isinstance(v_shares, str) and not v_shares.startswith('='):
            ws.cell(r, col_2026_shares).value = v_shares
            copy_style(ws.cell(r, col_2025_shares), ws.cell(r, col_2026_shares))
        if isinstance(v_amount, str) and not v_amount.startswith('='):
            ws.cell(r, col_2026_amount).value = v_amount
            copy_style(ws.cell(r, col_2025_amount), ws.cell(r, col_2026_amount))

        # 2026 No. of Shares col - formula if 2025 shares col has one, else blank
        if isinstance(v_shares, str) and v_shares.startswith('='):
            c26_sh = ws.cell(r, col_2026_shares)
            c26_sh.value = derive_2026_formula(v_shares, col_2025_shares, sheet_config)
            copy_style(ws.cell(r, col_2025_shares), c26_sh)
            c26_sh.fill = PatternFill("solid", fgColor="FFFBEB")
            c26_sh.font = Font(size=9, color="856404")

        # 2026 ₹ amount col - formula derived from 2025 ₹ col
        if isinstance(v_amount, str) and v_amount.startswith('='):
            c26_amt = ws.cell(r, col_2026_amount)
            c26_amt.value = derive_2026_formula(v_amount, col_2025_amount, sheet_config)
            copy_style(ws.cell(r, col_2025_amount), c26_amt)
            c26_amt.fill = PatternFill("solid", fgColor="FFFBEB")
            c26_amt.font = Font(size=9, color="856404")


# ══════════════════════════════════════════════════════════════
# PROCESS SUPPORT SHEET - exact v8 logic
# ══════════════════════════════════════════════════════════════

def process_support_sheet(ws, sheet_config):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and val.startswith('='):
                ws.cell(r, c).value = shift_formula(val, None, sheet_config)


# ══════════════════════════════════════════════════════════════
# NAMED RANGES - define TB named ranges & rewrite formulas
# ══════════════════════════════════════════════════════════════

def _build_row_to_name_map(named_ranges_cfg):
    """
    Build a lookup: (row, col_letter) -> name
    e.g. (175, 'B') -> 'TB_B_Advertisement_Exp'
    """
    row_map = {}
    for slug, (row_or_range, cols) in named_ranges_cfg.items():
        if isinstance(row_or_range, tuple):
            continue  # range names handled separately
        row = row_or_range
        for col in cols:          # 'B', 'C', or 'BC'
            name = f'TB_{col}_{slug}'
            row_map[(row, col)] = name
    return row_map


def create_tb_named_ranges(wb, named_ranges_cfg):
    """
    Define all TB Named Ranges in the workbook.
    Each name points to the exact cell in TB using an absolute reference.
    This runs ONCE on the output workbook - no row-number fragility after this.
    """
    if 'TB' not in wb.sheetnames:
        return

    for slug, (row_or_range, cols) in named_ranges_cfg.items():
        for col in cols:
            name = f'TB_{col}_{slug}'
            # Remove existing if any
            if name in wb.defined_names:
                del wb.defined_names[name]

            if isinstance(row_or_range, tuple):
                start_row, end_row = row_or_range
                ref = f"TB!${col}${start_row}:${col}${end_row}"
            else:
                row = row_or_range
                ref = f"TB!${col}${row}"

            wb.defined_names.add(DefinedName(name, attr_text=ref))


def rewrite_tb_formulas(wb, named_ranges_cfg, sheet_config):
    """
    Replace direct TB row/range references with their Named Range equivalents.
    Handles both single-cell refs (TB!B175) and range refs (TB!B126:B169).
    """
    # Build single-cell lookup: (row, col) -> name
    row_to_name = _build_row_to_name_map(named_ranges_cfg)

    # Build range lookup: (start_row, end_row, col) -> name
    range_to_name = {}
    for slug, (row_or_range, cols) in named_ranges_cfg.items():
        if isinstance(row_or_range, tuple):
            start_row, end_row = row_or_range
            for col in cols:
                name = f'TB_{col}_{slug}'
                range_to_name[(start_row, end_row, col)] = name

    # Pattern 1: range ref  TB!B126:B169 or TB!$B$126:$B$169
    TB_RANGE_PAT = re.compile(
        r'TB![$]?([BC])[$]?(\d+):[$]?([BC])[$]?(\d+)'
    )
    # Pattern 2: single-cell ref  TB!B175 or TB!$B$175
    # Only match when NOT followed by a colon (to avoid partial range matches)
    TB_CELL_PAT = re.compile(
        r'TB![$]?([BC])[$]?(\d+)(?!:\d|[$]\d)'
    )

    def replace_range_ref(m):
        col1, r1, col2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        name = range_to_name.get((r1, r2, col1))
        if name:
            return name
        return m.group(0)   # no named range - leave as-is

    def replace_cell_ref(m):
        col_ltr = m.group(1)
        row_num = int(m.group(2))
        name = row_to_name.get((row_num, col_ltr))
        if name:
            return name
        return m.group(0)   # no named range - leave as-is

    for shname in wb.sheetnames:
        ws = wb[shname]
        if shname == 'TB':
            continue   # don't rewrite the TB sheet itself
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str) or not v.startswith('=') or 'TB!' not in v:
                    continue
                # Replace range refs first, then single-cell refs
                new_v = TB_RANGE_PAT.sub(replace_range_ref, v)
                new_v = TB_CELL_PAT.sub(replace_cell_ref, new_v)
                if new_v != v:
                    ws.cell(r, c).value = new_v


# ══════════════════════════════════════════════════════════════
# CHAIN PATCHES - exact v8 logic
# ══════════════════════════════════════════════════════════════

def _apply_chain_patches(wb, sheet_config, chain_patches):
    for shname, r, c, orig_f in chain_patches:
        if shname not in wb.sheetnames: continue
        ins = sheet_config.get(shname, {}).get('insert')
        if ins:
            wb[shname].cell(r, c).value = shift_formula(orig_f, ins, sheet_config)
