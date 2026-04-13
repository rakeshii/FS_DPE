"""
core/column_detector.py
Multi-signal year column detection for financial statement sheets.

Instead of relying on hardcoded column indexes from config, this module
inspects each worksheet and identifies the current and preceding year
columns using weighted signals:

  1. Header text   — period marker with explicit year (weight 0.50)
  2. Numeric density — fraction of data rows with numbers  (weight 0.25)
  3. Formula density — fraction of data rows with formulas (weight 0.25)

Returns a ColumnDetectionResult with the detected columns and a confidence
score.  When confidence is too low, result.error is set instead of guessing.
"""

import re
from dataclasses import dataclass, field
from typing import List, Optional

from openpyxl.utils import get_column_letter

# Match the period labels used in the financial sheets
PERIOD_RE = re.compile(r'31\s+March|March\s+31|Year\s+Ended|31-03-|March,\s*20', re.I)
# Extract a 4-digit year (20xx) from a cell value
YEAR_RE = re.compile(r'(20\d{2})')

# Below this threshold the detector returns an error instead of a result
MIN_CONFIDENCE = 0.40

# How many header rows to scan for year labels
HEADER_SCAN_ROWS = 10


@dataclass
class ColumnSignal:
    """Per-column scoring data for the detection algorithm."""
    col_idx: int               # 1-indexed column number
    header_text: str = ''      # period label found in header rows
    header_year: int = 0       # year extracted from header (0 = not found)
    header_confidence: float = 0.0
    numeric_density: float = 0.0   # fraction of data rows that are numeric
    formula_density: float = 0.0   # fraction of data rows that are formulas
    is_label_col: bool = False     # True when column is mostly text labels
    total_score: float = 0.0


@dataclass
class ColumnDetectionResult:
    """Result returned by detect_year_columns()."""
    current_year_col: int          # 1-indexed: col to insert the new year BEFORE
    preceding_year_col: int        # 1-indexed: the column immediately to the right
    current_year: int              # detected year number (0 when unknown)
    confidence: float              # 0.0–1.0
    signals: List[ColumnSignal] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    error: Optional[str] = None    # set when detection fails or confidence too low


def detect_year_columns(ws, sheet_type: str = 'standard') -> ColumnDetectionResult:
    """
    Detect the current and preceding year columns in a financial worksheet.

    Algorithm
    ---------
    For each column in the sheet:
      1. Scan the first HEADER_SCAN_ROWS rows for a PERIOD_RE match.
         If found with a 4-digit year → header_confidence = 0.90.
         If found without a year      → header_confidence = 0.50.
      2. Count numeric and formula cells in all data rows (below header area).
      3. Classify columns that are mostly text as label columns (A, B, etc.).
      4. total_score = header_confidence×0.50 + numeric_density×0.25 + formula_density×0.25

    The column with the highest year number and the highest score is the
    'current year' column.  The next column to the right is 'preceding year'.

    Parameters
    ----------
    ws         : openpyxl Worksheet
    sheet_type : 'standard' | 'wide' | 'fixed_assets'  (for future specialisation)

    Returns
    -------
    ColumnDetectionResult
    """
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    data_start = HEADER_SCAN_ROWS + 1
    data_rows  = max(1, max_row - HEADER_SCAN_ROWS)

    signals: List[ColumnSignal] = []

    for col in range(1, max_col + 1):
        sig = ColumnSignal(col_idx=col)

        # ── Signal 1: header text scan ────────────────────────────────
        for r in range(1, HEADER_SCAN_ROWS + 1):
            raw = ws.cell(r, col).value
            if raw is None:
                continue
            text = str(raw)
            if PERIOD_RE.search(text):
                sig.header_text = text.strip()
                m = YEAR_RE.search(text)
                if m:
                    sig.header_year       = int(m.group(1))
                    sig.header_confidence = 0.90
                else:
                    sig.header_confidence = 0.50
                break  # first match per column is enough

        # ── Signal 2+3: numeric / formula density ─────────────────────
        numeric_count = 0
        formula_count = 0
        for r in range(data_start, max_row + 1):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                numeric_count += 1
            elif isinstance(v, str) and v.startswith('='):
                formula_count += 1

        sig.numeric_density = numeric_count / data_rows
        sig.formula_density = formula_count / data_rows

        # ── Signal 4: label column detection ─────────────────────────
        # A column is a label column if it has many non-formula text cells
        # and almost no numbers or formulas (typical for columns A and B).
        text_count = 0
        check_up_to = min(max_row + 1, data_start + 50)
        for r in range(data_start, check_up_to):
            v = ws.cell(r, col).value
            if isinstance(v, str) and not v.startswith('=') and len(v.strip()) > 2:
                text_count += 1
        sig.is_label_col = (
            text_count >= 8 and
            sig.numeric_density < 0.10 and
            sig.formula_density < 0.10
        )

        # ── Weighted total score ──────────────────────────────────────
        if sig.is_label_col:
            sig.total_score = 0.0
        else:
            sig.total_score = (
                sig.header_confidence * 0.50 +
                sig.numeric_density   * 0.25 +
                sig.formula_density   * 0.25
            )

        signals.append(sig)

    # ── Pick current year column ──────────────────────────────────────
    year_cols = [s for s in signals if s.header_year > 0 and not s.is_label_col]

    if not year_cols:
        # Fallback: use numeric/formula density alone (lower confidence)
        data_dense = sorted(
            [s for s in signals
             if not s.is_label_col and (s.numeric_density + s.formula_density) > 0.05],
            key=lambda s: s.col_idx
        )
        if data_dense:
            cur = data_dense[0]
            return ColumnDetectionResult(
                current_year_col   = cur.col_idx,
                preceding_year_col = cur.col_idx + 1,
                current_year       = 0,
                confidence         = 0.20,
                signals            = signals,
                warnings = [
                    'No year header found in this sheet; '
                    'current year column was inferred from numeric density only. '
                    'Review the output carefully.'
                ]
            )
        # Complete failure
        return ColumnDetectionResult(
            current_year_col   = 4,
            preceding_year_col = 5,
            current_year       = 0,
            confidence         = 0.0,
            signals            = signals,
            error = (
                f"Could not detect any year-labeled columns in this sheet.\n"
                f"Expected cells containing '31 March YYYY' or 'Year Ended YYYY' "
                f"in the first {HEADER_SCAN_ROWS} rows.\n"
                f"Sheet dimensions: {max_col} columns × {max_row} rows.\n"
                "Verify the uploaded file is the correct financial statements workbook."
            )
        )

    # Highest year found → that is the current year to be frozen / insert before
    max_year = max(s.header_year for s in year_cols)
    candidates = sorted(
        [s for s in year_cols if s.header_year == max_year],
        key=lambda s: s.col_idx
    )
    current    = candidates[0]   # leftmost column with the current year
    confidence = current.total_score

    warnings: List[str] = []
    if len(candidates) > 1:
        warnings.append(
            f"Multiple columns found with year {max_year}: "
            f"cols {[get_column_letter(s.col_idx) for s in candidates]}. "
            f"Using leftmost ({get_column_letter(current.col_idx)})."
        )

    if confidence < MIN_CONFIDENCE:
        return ColumnDetectionResult(
            current_year_col   = current.col_idx,
            preceding_year_col = current.col_idx + 1,
            current_year       = max_year,
            confidence         = confidence,
            signals            = signals,
            warnings           = warnings,
            error = (
                f"Low detection confidence ({confidence:.0%}) for current year column "
                f"{get_column_letter(current.col_idx)} "
                f"(header: '{current.header_text}').\n"
                "Possible causes:\n"
                "  • The year column has very few numeric or formula cells\n"
                "  • Header format differs from 'As at 31 March, YYYY'\n"
                "  • Sheet layout has changed significantly from the expected template\n"
                "Update SHEET_CONFIG['insert'] in config.py if the column position changed."
            )
        )

    return ColumnDetectionResult(
        current_year_col   = current.col_idx,
        preceding_year_col = current.col_idx + 1,
        current_year       = max_year,
        confidence         = confidence,
        signals            = signals,
        warnings           = warnings,
    )
