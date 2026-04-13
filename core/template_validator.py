"""
core/template_validator.py
TemplateValidator: profile checks for uploaded financial statements workbooks.

Instead of partial output on structural mismatches, this module returns
structured ValidationIssue objects so the caller can decide whether to
abort or proceed with warnings.

Checks performed
----------------
1. Required sheet presence   — BS, P&L, CFS must exist (error if missing)
2. Period header presence     — each configured sheet must have a year-end
                               label in the first 15 rows (error if absent)
3. Column detection          — year column must be detectable with acceptable
                               confidence (error for core sheets, warning for notes)
4. Config vs detected mismatch — when detected column differs from config insert
                               (warning, never error — config may be intentional)
"""

import re
from dataclasses import dataclass, field
from typing import List, Optional

from openpyxl.utils import get_column_letter

from core.column_detector import detect_year_columns, ColumnDetectionResult

# Must match the period labels used in financial sheets
PERIOD_RE = re.compile(r'31\s+March|March\s+31|Year\s+Ended|31-03-|March,\s*20', re.I)

# These three sheets are mandatory; detection failures on them are errors
CORE_SHEETS = frozenset({'BS', 'P&L', 'CFS'})


@dataclass
class ValidationIssue:
    severity: str            # 'error' | 'warning'
    sheet: str
    row: Optional[int]
    col: Optional[int]       # 1-indexed; None when not applicable
    reason: str
    suggestion: str = ''

    def __str__(self) -> str:
        loc = f"[{self.sheet}"
        if self.row is not None:
            loc += f" row {self.row}"
        if self.col is not None:
            loc += f" col {get_column_letter(self.col)}"
        loc += "]"
        msg = f"{self.severity.upper()} {loc}: {self.reason}"
        if self.suggestion:
            msg += f"\n   → {self.suggestion}"
        return msg


@dataclass
class TemplateValidationResult:
    passed: bool
    issues: List[ValidationIssue] = field(default_factory=list)
    # sheet name → ColumnDetectionResult for every processed sheet
    detected_columns: dict = field(default_factory=dict)

    def errors(self) -> List[ValidationIssue]:
        return [i for i in self.issues if i.severity == 'error']

    def warnings(self) -> List[ValidationIssue]:
        return [i for i in self.issues if i.severity == 'warning']

    def format_report(self) -> str:
        lines: List[str] = []
        errs  = self.errors()
        warns = self.warnings()
        if errs:
            lines.append(f"VALIDATION FAILED — {len(errs)} error(s):")
            for e in errs:
                lines.append(f"  • {e}")
        if warns:
            if lines:
                lines.append('')
            lines.append(f"Warnings ({len(warns)}):")
            for w in warns:
                lines.append(f"  ⚠  {w}")
        if not errs and not warns:
            lines.append("All template validation checks passed.")
        return '\n'.join(lines)


class TemplateValidator:
    """
    Validate an openpyxl Workbook against the expected FS template profile.

    Usage
    -----
    validator = TemplateValidator(sheet_config)
    result = validator.validate(wb)
    if not result.passed:
        raise RuntimeError(result.format_report())
    """

    def __init__(self, sheet_config: dict):
        self.sheet_config = sheet_config

    def validate(self, wb) -> TemplateValidationResult:
        issues: List[ValidationIssue] = []
        detected: dict = {}
        sheet_names = set(wb.sheetnames)

        # ── 1. Required sheet presence ────────────────────────────────
        for name in sorted(CORE_SHEETS - sheet_names):
            issues.append(ValidationIssue(
                severity   = 'error',
                sheet      = name,
                row        = None,
                col        = None,
                reason     = f"Required sheet '{name}' is missing from the workbook.",
                suggestion = (
                    "Upload the complete financial statements file that contains "
                    "Balance Sheet (BS), Profit & Loss (P&L), and Cash Flow (CFS) sheets."
                )
            ))

        # If core sheets are absent there is nothing more to check
        if any(i.severity == 'error' for i in issues):
            return TemplateValidationResult(
                passed=False, issues=issues, detected_columns=detected
            )

        # ── 2. Per-sheet checks ───────────────────────────────────────
        for shname, cfg in self.sheet_config.items():
            if shname not in sheet_names:
                issues.append(ValidationIssue(
                    severity   = 'warning',
                    sheet      = shname,
                    row        = None,
                    col        = None,
                    reason     = (
                        f"Configured sheet '{shname}' is not present in the workbook "
                        "and will be skipped."
                    ),
                    suggestion = (
                        "If the sheet was renamed in the new template, update "
                        "SHEET_CONFIG in config.py to match."
                    )
                ))
                continue

            ws         = wb[shname]
            sheet_type = cfg.get('type', 'standard')
            cfg_insert = cfg.get('insert')
            is_core    = shname in CORE_SHEETS

            # ── 2a. Period header presence ──────────────────────────
            period_found = False
            for r in range(1, min(16, ws.max_row + 1)):
                for c in range(1, ws.max_column + 1):
                    if PERIOD_RE.search(str(ws.cell(r, c).value or '')):
                        period_found = True
                        break
                if period_found:
                    break

            if not period_found:
                issues.append(ValidationIssue(
                    severity   = 'error' if is_core else 'warning',
                    sheet      = shname,
                    row        = None,
                    col        = None,
                    reason     = (
                        "No period header (e.g. 'As at 31 March, 2025') found "
                        "in the first 15 rows of this sheet."
                    ),
                    suggestion = (
                        "Verify this is the correct sheet and that the year-end date "
                        "label is present in the header area (rows 1–15)."
                    )
                ))

            # ── 2b. Column detection ────────────────────────────────
            detection = detect_year_columns(ws, sheet_type=sheet_type)
            detected[shname] = detection

            if detection.error:
                # Detection errors on core sheets block processing;
                # on note sheets they are downgraded to warnings.
                issues.append(ValidationIssue(
                    severity   = 'error' if is_core else 'warning',
                    sheet      = shname,
                    row        = None,
                    col        = detection.current_year_col,
                    reason     = detection.error,
                    suggestion = (
                        "Check that this sheet has year-date column headers and "
                        "financial data in the expected layout."
                    )
                ))
            else:
                for w in detection.warnings:
                    issues.append(ValidationIssue(
                        severity = 'warning',
                        sheet    = shname,
                        row      = None,
                        col      = detection.current_year_col,
                        reason   = w
                    ))

            # ── 2c. Config vs detected column mismatch ──────────────
            if (cfg_insert and not detection.error and
                    detection.current_year_col != cfg_insert):
                issues.append(ValidationIssue(
                    severity   = 'warning',
                    sheet      = shname,
                    row        = None,
                    col        = detection.current_year_col,
                    reason     = (
                        f"Detected current year in column "
                        f"{get_column_letter(detection.current_year_col)} "
                        f"(index {detection.current_year_col}), but config.py "
                        f"specifies insert={cfg_insert} "
                        f"({get_column_letter(cfg_insert)})."
                    ),
                    suggestion = (
                        f"If the template column layout changed, update "
                        f"SHEET_CONFIG['{shname}']['insert'] to "
                        f"{detection.current_year_col} in config.py."
                    )
                ))

        has_errors = any(i.severity == 'error' for i in issues)
        return TemplateValidationResult(
            passed           = not has_errors,
            issues           = issues,
            detected_columns = detected
        )
